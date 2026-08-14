import os
import openpyxl
import re
import uuid
from datetime import datetime, date, timedelta
from backend.storage import EXCEL_FILE

EXCEL_FILE = str(EXCEL_FILE)

COURSE_HEADERS = ['Plan', 'Course Name', 'Path', 'Module Name', 'Duration', 'Duration (Minutes)', 'Queue', '% of Total', 'Module ID', 'Course Order', 'Module Order', 'Source Module ID']
PLAN_META_SHEET = "Plan Metadata"
PLAN_META_HEADERS = ['Plan ID', 'Plan Name', 'Plan Type', 'Source Plan ID', 'Display Order', 'Version', 'Updated At']

def _stable_id(prefix, *parts):
    raw = "|".join(str(p or "").strip() for p in parts)
    return f"{prefix}_{uuid.uuid5(uuid.NAMESPACE_URL, raw).hex[:16]}"

def _ensure_plan_metadata(wb):
    """Migrate legacy name-based plans to stable metadata without changing row order."""
    dirty = False
    course_ws = next(ws for ws in wb.worksheets if ws.cell(1, 1).value == "Plan" and ws.cell(1, 2).value == "Course Name")
    for col, header in enumerate(COURSE_HEADERS, 1):
        if course_ws.cell(1, col).value != header:
            course_ws.cell(1, col).value = header
            dirty = True

    if PLAN_META_SHEET not in wb.sheetnames:
        meta_ws = wb.create_sheet(PLAN_META_SHEET)
        meta_ws.append(PLAN_META_HEADERS)
        dirty = True
    else:
        meta_ws = wb[PLAN_META_SHEET]
        for col, header in enumerate(PLAN_META_HEADERS, 1):
            if meta_ws.cell(1, col).value != header:
                meta_ws.cell(1, col).value = header
                dirty = True

    existing = {}
    for r in range(2, meta_ws.max_row + 1):
        name = str(meta_ws.cell(r, 2).value or "").strip()
        if name:
            existing[name] = r

    plan_order = []
    course_orders = {}
    module_orders = {}
    for r in range(2, course_ws.max_row + 1):
        plan = str(course_ws.cell(r, 1).value or "").strip()
        course = str(course_ws.cell(r, 2).value or "").strip()
        path = str(course_ws.cell(r, 3).value or "").strip()
        module = str(course_ws.cell(r, 4).value or "").strip()
        if not plan or not course:
            continue
        if plan not in plan_order:
            plan_order.append(plan)
        course_key = (plan, course, path)
        if course_key not in course_orders:
            course_orders[course_key] = 1 + sum(1 for k in course_orders if k[0] == plan)
        module_orders[course_key] = module_orders.get(course_key, 0) + 1
        if not course_ws.cell(r, 9).value:
            course_ws.cell(r, 9).value = _stable_id("mod", plan, course, path, module)
            dirty = True
        if course_ws.cell(r, 10).value is None:
            course_ws.cell(r, 10).value = course_orders[course_key]
            dirty = True
        if course_ws.cell(r, 11).value is None:
            course_ws.cell(r, 11).value = module_orders[course_key]
            dirty = True

    for idx, plan in enumerate(plan_order, 1):
        if plan not in existing:
            # Legacy copied plans are treated as templates when their name starts
            # with an existing plan name; future clones store this explicitly.
            source = next((p for p in plan_order if p != plan and plan.startswith(p + " -")), None)
            source_id = _stable_id("plan", source) if source else ""
            meta_ws.append([_stable_id("plan", plan), plan, "template" if source else "main", source_id, idx, 1, datetime.now()])
            dirty = True
    return dirty

def init_db():
    """Ensure necessary sheets exist in sheet.xlsx with headers."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách khóa học EPM V5"
        # Write headers
        ws.append(COURSE_HEADERS)
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    dirty = False

    # 1. Check Course sheet
    if "Danh sách khóa học EPM V5" not in wb.sheetnames:
        ws = wb.create_sheet("Danh sách khóa học EPM V5")
        ws.append(COURSE_HEADERS)
        dirty = True

    dirty = _ensure_plan_metadata(wb) or dirty

    # 2. Check Employees sheet
    if "Danh sách nhân viên" not in wb.sheetnames:
        ws = wb.create_sheet("Danh sách nhân viên")
        headers = ['Username', 'Full Name', 'English Name', 'Role', 'Email', 'PasswordHash', 'MustChangePassword', 'AvatarUrl']
        ws.append(headers)
        dirty = True
    else:
        ws = wb["Danh sách nhân viên"]
        headers = ['Username', 'Full Name', 'English Name', 'Role', 'Email', 'PasswordHash', 'MustChangePassword', 'AvatarUrl']
        for col_idx, h in enumerate(headers, 1):
            if ws.cell(row=1, column=col_idx).value != h:
                ws.cell(row=1, column=col_idx).value = h
                dirty = True
        
        # Populate initial hashed passwords for existing employees if empty
        default_pwd_hash = hash_password(DEFAULT_INITIAL_PASSWORD)
        for r in range(2, ws.max_row + 1):
            u_val = ws.cell(row=r, column=1).value
            pwd_val = ws.cell(row=r, column=6).value
            must_change = ws.cell(row=r, column=7).value
            if u_val is not None and str(u_val).strip() != "":
                if not pwd_val:
                    ws.cell(row=r, column=6).value = default_pwd_hash
                    ws.cell(row=r, column=7).value = False
                    dirty = True
                elif must_change is None:
                    ws.cell(row=r, column=7).value = False
                    dirty = True

    # 3. Check Progress sheet
    if "Tiến độ học tập" not in wb.sheetnames:
        ws = wb.create_sheet("Tiến độ học tập")
        headers = [
            'Username', 'Course Name', 'Path', 'Module Name', 'Status', 
            'Progress (%)', 'Start Date', 'Completion Date', 'Planned Completion Date', 'Tracking Status',
            'Planned Start Date', 'Module ID', 'Enrollment ID', 'Excluded',
            'Planned Start Offset Minutes', 'Planned End Offset Minutes'
        ]
        ws.append(headers)
        dirty = True
    else:
        ws = next(s for s in wb.worksheets if s.cell(1, 1).value == 'Username' and s.cell(1, 5).value == 'Status')
        headers = ['Username', 'Course Name', 'Path', 'Module Name', 'Status', 'Progress (%)', 'Start Date', 'Completion Date', 'Planned Completion Date', 'Tracking Status', 'Planned Start Date', 'Module ID', 'Enrollment ID', 'Excluded', 'Planned Start Offset Minutes', 'Planned End Offset Minutes']
        for col_idx, header in enumerate(headers, 1):
            if ws.cell(1, col_idx).value != header:
                ws.cell(1, col_idx).value = header
                dirty = True

    # 4. Check Enrollments sheet
    if "Đăng ký khóa học" not in wb.sheetnames:
        ws = wb.create_sheet("Đăng ký khóa học")
        ws.append(['Username', 'Target Type', 'Target Name', 'Start Date', 'Planned End Date', 'Workweek Type', 'Ratio', 'Daily Hours', 'Enrollment ID', 'Target ID', 'Target Version'])
        dirty = True

    enrollment_ws = next((s for s in wb.worksheets if s.cell(1, 1).value == 'Username' and s.cell(1, 2).value == 'Target Type'), None)
    if enrollment_ws is not None:
        enrollment_headers = ['Username', 'Target Type', 'Target Name', 'Start Date', 'Planned End Date', 'Workweek Type', 'Ratio', 'Daily Hours', 'Enrollment ID', 'Target ID', 'Target Version']
        for col_idx, header in enumerate(enrollment_headers, 1):
            if enrollment_ws.cell(1, col_idx).value != header:
                enrollment_ws.cell(1, col_idx).value = header
                dirty = True
        meta_ws = wb[PLAN_META_SHEET]
        meta_by_name = {str(meta_ws.cell(r, 2).value): (meta_ws.cell(r, 1).value, meta_ws.cell(r, 6).value or 1) for r in range(2, meta_ws.max_row + 1)}
        for r in range(2, enrollment_ws.max_row + 1):
            username = str(enrollment_ws.cell(r, 1).value or '').strip()
            target_name = str(enrollment_ws.cell(r, 3).value or '').strip()
            if username and not enrollment_ws.cell(r, 9).value:
                enrollment_ws.cell(r, 9).value = _stable_id('enr', username, enrollment_ws.cell(r, 2).value, target_name)
                dirty = True
            if target_name in meta_by_name:
                target_id, version = meta_by_name[target_name]
                if not enrollment_ws.cell(r, 10).value:
                    enrollment_ws.cell(r, 10).value = target_id
                    dirty = True
                if enrollment_ws.cell(r, 11).value is None:
                    enrollment_ws.cell(r, 11).value = version
                    dirty = True

    if dirty:
        wb.save(EXCEL_FILE)
    wb.close()

def get_courses():
    """Retrieve all course modules."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Danh sách khóa học EPM V5"]
    meta_ws = wb[PLAN_META_SHEET]
    plan_meta = {}
    for mr in range(2, meta_ws.max_row + 1):
        plan_meta[str(meta_ws.cell(mr, 2).value or '').strip()] = {
            "plan_id": meta_ws.cell(mr, 1).value,
            "plan_type": meta_ws.cell(mr, 3).value or "main",
            "source_plan_id": meta_ws.cell(mr, 4).value or "",
            "plan_order": int(meta_ws.cell(mr, 5).value or mr - 1),
            "plan_version": int(meta_ws.cell(mr, 6).value or 1)
        }
    courses = []
    # Headers are: Plan, Course Name, Path, Module Name, Duration, Duration (Minutes), Queue, % of Total
    for r in range(2, ws.max_row + 1):
        plan = ws.cell(row=r, column=1).value
        course_name = ws.cell(row=r, column=2).value
        path = ws.cell(row=r, column=3).value
        module_name = ws.cell(row=r, column=4).value
        duration = ws.cell(row=r, column=5).value
        duration_mins = ws.cell(row=r, column=6).value
        queue = ws.cell(row=r, column=7).value
        pct = ws.cell(row=r, column=8).value
        module_id = ws.cell(row=r, column=9).value
        course_order = ws.cell(row=r, column=10).value
        module_order = ws.cell(row=r, column=11).value
        source_module_id = ws.cell(row=r, column=12).value
        
        # Normalize Queue (can be boolean, string, or cell formula representation)
        is_queue = False
        if queue is not None:
            if isinstance(queue, bool):
                is_queue = queue
            elif str(queue).strip().upper() in ("TRUE", "1", "YES"):
                is_queue = True
        
        if course_name:
            meta = plan_meta.get(str(plan or '').strip(), {})
            courses.append({
                "row": r,
                "plan": plan,
                "course_name": course_name,
                "path": path if path else "",
                "module_name": module_name if module_name else "",
                "duration": duration if duration else "",
                "duration_minutes": int(duration_mins) if duration_mins is not None else 0,
                "queue": is_queue,
                "percent_of_total": pct if pct else "0.00%",
                "module_id": module_id or _stable_id("mod", plan, course_name, path, module_name),
                "course_id": _stable_id("course", meta.get("plan_id", plan), course_name, path),
                "source_module_id": source_module_id or "",
                "course_order": int(course_order or 0),
                "module_order": int(module_order or 0),
                **meta
            })
    wb.close()
    courses.sort(key=lambda c: (c.get("plan_order", 999999), c.get("course_order", 999999), c.get("module_order", 999999), c["row"]))
    return courses

def recalculate_formulas(ws):
    """Recalculate % of Total formulas in column H."""
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        # Format formula: =IF(G{row}, TEXT(F{row}/SUMIF($G$2:$G${max_row}, TRUE, $F$2:$F${max_row}), "0.00%"), "0.00%")
        formula = f'=IF(G{r}, TEXT(F{r}/SUMIF($G$2:$G${max_row}, TRUE, $F$2:$F${max_row}), "0.00%"), "0.00%")'
        ws.cell(row=r, column=8).value = formula

def add_course_modules(plan, course_name, path, modules):
    """
    modules is a list of dicts: [{"module_name": str, "duration": str, "duration_minutes": int, "queue": bool}]
    Inserts or appends modules under the given course and path.
    """
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách khóa học EPM V5"]
    
    existing_course_orders = [int(ws.cell(r, 10).value or 0) for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == plan]
    course_order = next((int(ws.cell(r, 10).value or 0) for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == plan and ws.cell(r, 2).value == course_name and (ws.cell(r, 3).value or "") == (path or "")), max(existing_course_orders or [0]) + 1)
    existing_module_orders = [int(ws.cell(r, 11).value or 0) for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == plan and ws.cell(r, 2).value == course_name and (ws.cell(r, 3).value or "") == (path or "")]
    next_module_order = max(existing_module_orders or [0]) + 1

    # Find existing row for this course
    insert_at = None
    for r in range(2, ws.max_row + 1):
        plan_name = ws.cell(row=r, column=1).value
        c_name = ws.cell(row=r, column=2).value
        p_name = ws.cell(row=r, column=3).value
        if plan_name == plan and c_name == course_name and (not path or p_name == path):
            insert_at = r
            # Find the last module of this course/path
            while r <= ws.max_row and ws.cell(row=r, column=1).value == plan and ws.cell(row=r, column=2).value == course_name and (not path or ws.cell(row=r, column=3).value == path):
                insert_at = r
                r += 1
            break
            
    if insert_at is None:
        # Append at the end
        for m in modules:
            module_id = m.get("module_id") or _stable_id("mod", plan, course_name, path, m["module_name"], uuid.uuid4().hex)
            ws.append([
                plan, course_name, path or "", m["module_name"], 
                m["duration"], m["duration_minutes"], m.get("queue", True), "",
                module_id, course_order, next_module_order, m.get("source_module_id", "")
            ])
            next_module_order += 1
    else:
        # Insert rows right after the last module of the course
        for i, m in enumerate(modules):
            idx = insert_at + 1 + i
            ws.insert_rows(idx)
            ws.cell(row=idx, column=1).value = plan
            ws.cell(row=idx, column=2).value = course_name
            ws.cell(row=idx, column=3).value = path or ""
            ws.cell(row=idx, column=4).value = m["module_name"]
            ws.cell(row=idx, column=5).value = m["duration"]
            ws.cell(row=idx, column=6).value = m["duration_minutes"]
            ws.cell(row=idx, column=7).value = m.get("queue", True)
            ws.cell(row=idx, column=9).value = m.get("module_id") or _stable_id("mod", plan, course_name, path, m["module_name"], uuid.uuid4().hex)
            ws.cell(row=idx, column=10).value = course_order
            ws.cell(row=idx, column=11).value = next_module_order
            ws.cell(row=idx, column=12).value = m.get("source_module_id", "")
            next_module_order += 1
            
    recalculate_formulas(ws)
    wb.save(EXCEL_FILE)
    wb.close()
    _touch_plan_and_reschedule(plan)

def delete_course(course_name):
    """Delete all rows matching course_name."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách khóa học EPM V5"]
    
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == course_name:
            rows_to_delete.append(r)
            
    # Delete from bottom up
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)
        
    recalculate_formulas(ws)
    wb.save(EXCEL_FILE)
    wb.close()

def toggle_module_status(plan: Optional[str] = None, course_name: Optional[str] = None, module_name: Optional[str] = None, path: Optional[str] = None, queue: bool = True):
    """Toggle Queue (active/inactive) status for a specific module, path, or all modules in a course."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách khóa học EPM V5"]
    
    updated = False
    for r in range(2, ws.max_row + 1):
        p_val = ws.cell(row=r, column=1).value
        c_val = ws.cell(row=r, column=2).value
        path_val = ws.cell(row=r, column=3).value
        m_val = ws.cell(row=r, column=4).value
        
        match_plan = (p_val == plan) if plan else True
        match_course = (c_val == course_name) if course_name else True
        match_path = (path_val == path) if path is not None else True
        match_module = (m_val == module_name) if module_name else True
        
        if match_plan and match_course and match_path and match_module:
            ws.cell(row=r, column=7).value = queue
            updated = True
            
    if updated:
        recalculate_formulas(ws)
        wb.save(EXCEL_FILE)
    wb.close()
    if updated and plan:
        _touch_plan_and_reschedule(plan)
    return updated

def update_module_duration(plan: Optional[str], course_name: str, module_name: str, duration: str, duration_minutes: int, path: Optional[str] = None):
    """Update duration string and duration_minutes for a specific module in sheet.xlsx."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách khóa học EPM V5"]
    
    updated = False
    for r in range(2, ws.max_row + 1):
        p_val = ws.cell(row=r, column=1).value
        c_val = ws.cell(row=r, column=2).value
        path_val = ws.cell(row=r, column=3).value
        m_val = ws.cell(row=r, column=4).value
        
        match_plan = (p_val == plan) if plan else True
        match_course = (c_val == course_name)
        match_path = (path_val == path) if path is not None else True
        match_module = (m_val == module_name)
        
        if match_plan and match_course and match_path and match_module:
            ws.cell(row=r, column=5).value = str(duration)
            ws.cell(row=r, column=6).value = int(duration_minutes)
            updated = True
            
    if updated:
        recalculate_formulas(ws)
        wb.save(EXCEL_FILE)
    wb.close()
    if updated and plan:
        _touch_plan_and_reschedule(plan)
    return updated

def move_module_to_course(source_plan: Optional[str], source_course: str, module_name: str, target_plan: str, target_course: str, target_path: Optional[str] = None, source_path: Optional[str] = None):
    """Move a module from source course to target course in sheet.xlsx."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách khóa học EPM V5"]
    
    moved = False
    for r in range(2, ws.max_row + 1):
        p_val = ws.cell(row=r, column=1).value
        c_val = ws.cell(row=r, column=2).value
        path_val = ws.cell(row=r, column=3).value
        m_val = ws.cell(row=r, column=4).value
        
        match_plan = (p_val == source_plan) if source_plan else True
        match_course = (c_val == source_course)
        match_path = (path_val == source_path) if source_path is not None else True
        match_module = (m_val == module_name)
        
        if match_plan and match_course and match_path and match_module:
            ws.cell(row=r, column=1).value = target_plan
            ws.cell(row=r, column=2).value = target_course
            ws.cell(row=r, column=3).value = target_path if target_path is not None else path_val
            moved = True
            break
            
    if moved:
        recalculate_formulas(ws)
        wb.save(EXCEL_FILE)
    wb.close()
    if moved:
        if source_plan:
            _touch_plan_and_reschedule(source_plan)
        _touch_plan_and_reschedule(target_plan)
    return moved

def delete_single_module(plan: Optional[str], course_name: str, module_name: str, path: Optional[str] = None):
    """Delete a single module row from sheet.xlsx."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách khóa học EPM V5"]
    
    target_row = None
    for r in range(2, ws.max_row + 1):
        p_val = ws.cell(row=r, column=1).value
        c_val = ws.cell(row=r, column=2).value
        path_val = ws.cell(row=r, column=3).value
        m_val = ws.cell(row=r, column=4).value
        
        match_plan = (p_val == plan) if plan else True
        match_course = (c_val == course_name)
        match_path = (path_val == path) if path is not None else True
        match_module = (m_val == module_name)
        
        if match_plan and match_course and match_path and match_module:
            target_row = r
            break
            
    if target_row:
        ws.delete_rows(target_row)
        recalculate_formulas(ws)
        wb.save(EXCEL_FILE)
        wb.close()
        if plan:
            _touch_plan_and_reschedule(plan)
        return True
        
    wb.close()
    return False

def clone_course_campaign(source_course_name: str, new_course_name: str, new_plan: Optional[str] = None, new_path: Optional[str] = None, source_path: Optional[str] = None):
    """Clone active modules of a course into a new independent course template / campaign."""
    init_db()
    all_courses = get_courses()
    
    # Filter active modules from source course
    matching = [
        c for c in all_courses 
        if c["course_name"] == source_course_name 
        and (not source_path or c["path"] == source_path)
        and c.get("queue", True)
    ]
    
    if not matching:
        return False
        
    target_plan = new_plan or matching[0]["plan"]
    target_path = new_path if new_path is not None else matching[0]["path"]
    
    modules_to_add = []
    for m in matching:
        modules_to_add.append({
            "module_name": m["module_name"],
            "duration": m["duration"],
            "duration_minutes": m["duration_minutes"],
            "queue": True
        })
        
    add_course_modules(target_plan, new_course_name, target_path, modules_to_add)
    return True

def clone_plan_campaign(source_plan: str, new_plan: str):
    """Clone all active courses and modules of a source Plan into a new independent Plan."""
    init_db()
    all_courses = get_courses()
    
    matching = [
        c for c in all_courses 
        if c["plan"] == source_plan 
        and c.get("queue", True)
    ]
    
    if not matching:
        return False

    wb = openpyxl.load_workbook(EXCEL_FILE)
    meta_ws = wb[PLAN_META_SHEET]
    meta_by_name = {str(meta_ws.cell(r, 2).value or '').strip(): r for r in range(2, meta_ws.max_row + 1)}
    if new_plan in meta_by_name:
        wb.close()
        return False
    source_row = meta_by_name.get(source_plan)
    source_id = meta_ws.cell(source_row, 1).value if source_row else _stable_id('plan', source_plan)
    max_order = max([int(meta_ws.cell(r, 5).value or 0) for r in range(2, meta_ws.max_row + 1)] or [0])
    meta_ws.append([_stable_id('plan', new_plan), new_plan, 'template', source_id, max_order + 1, 1, datetime.now()])
    wb.save(EXCEL_FILE)
    wb.close()
        
    course_groups = {}
    for m in matching:
        key = (m["course_name"], m["path"])
        if key not in course_groups:
            course_groups[key] = []
        course_groups[key].append({
            "module_name": m["module_name"],
            "duration": m["duration"],
            "duration_minutes": m["duration_minutes"],
            "queue": True,
            "source_module_id": m.get("module_id", "")
        })
        
    for (course_name, path), modules in course_groups.items():
        add_course_modules(new_plan, course_name, path, modules)
        
    return True

def add_source_module_to_template(template_plan: str, source_module_id: str):
    """Include one module from a template's linked main Plan."""
    init_db()
    courses = get_courses()
    template_modules = [m for m in courses if m['plan'] == template_plan]
    if not template_modules:
        return False
    template_meta = _get_plan_meta_by_name(template_plan)
    if template_meta.get('plan_type') != 'template' or not template_meta.get('source_plan_id'):
        return False
    source = next((m for m in courses if m.get('module_id') == source_module_id and m.get('plan_id') == template_meta['source_plan_id']), None)
    if not source:
        return False
    existing = next((m for m in template_modules if m.get('source_module_id') == source_module_id), None)
    if existing:
        if not existing.get('queue'):
            return toggle_module_status(template_plan, existing['course_name'], existing['module_name'], existing['path'], True)
        return True
    add_course_modules(template_plan, source['course_name'], source['path'], [{
        'module_name': source['module_name'],
        'duration': source['duration'],
        'duration_minutes': source['duration_minutes'],
        'queue': True,
        'source_module_id': source_module_id
    }])
    return True

def reorder_courses_in_plan(plan: str, course_order: List[str]):
    """Reorder course row blocks in sheet.xlsx for a given plan according to course_order list."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách khóa học EPM V5"]
    
    all_other_rows = []
    plan_rows = []
    plan_insert_index = None
    
    for r in range(2, ws.max_row + 1):
        p_val = ws.cell(row=r, column=1).value
        row_values = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        if p_val == plan:
            if plan_insert_index is None:
                plan_insert_index = len(all_other_rows)
            plan_rows.append(row_values)
        else:
            all_other_rows.append(row_values)
            
    if not plan_rows:
        wb.close()
        return False

    # Group plan_rows by course_name
    grouped = {}
    for row in plan_rows:
        c_name = row[1]  # Col B: Course Name
        if c_name not in grouped:
            grouped[c_name] = []
        grouped[c_name].append(row)
        
    # Reassemble plan rows according to course_order
    reordered_plan_rows = []
    for order_idx, c_name in enumerate(course_order, 1):
        if c_name in grouped:
            course_rows = grouped.pop(c_name)
            for row in course_rows:
                row[9] = order_idx
            reordered_plan_rows.extend(course_rows)
            
    # Append any remaining courses in plan not explicitly listed
    next_order = len(course_order) + 1
    for remaining_rows in grouped.values():
        for row in remaining_rows:
            row[9] = next_order
        reordered_plan_rows.extend(remaining_rows)
        next_order += 1
        
    # Combine non-plan rows and reordered plan rows
    # Reinsert the reordered block at its original position so editing a Plan
    # never moves the Plan card to the bottom of the catalog.
    insert_at = plan_insert_index if plan_insert_index is not None else len(all_other_rows)
    final_rows = all_other_rows[:insert_at] + reordered_plan_rows + all_other_rows[insert_at:]
    
    # Clear existing worksheet rows from row 2
    ws.delete_rows(2, ws.max_row)
    
    # Write back final_rows
    for row_idx, rdata in enumerate(final_rows, start=2):
        for col_idx, val in enumerate(rdata, start=1):
            ws.cell(row=row_idx, column=col_idx).value = val
            
    recalculate_formulas(ws)
    wb.save(EXCEL_FILE)
    wb.close()
    _touch_plan_and_reschedule(plan)
    return True

import hashlib
import hmac

DEFAULT_INITIAL_PASSWORD = "EPM@2026"

def hash_password(password: str, salt: str = None) -> str:
    """Hash password using PBKDF2_HMAC SHA256."""
    if not password:
        password = DEFAULT_INITIAL_PASSWORD
    if not salt:
        salt = os.urandom(16).hex()
    key = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000)
    return f"{salt}${key.hex()}"

def verify_password(stored_password_hash: str, password_candidate: str) -> bool:
    """Verify candidate password against stored PBKDF2 hash."""
    if not stored_password_hash:
        return False
    if '$' not in stored_password_hash:
        return stored_password_hash == password_candidate
    salt, hashed_hex = stored_password_hash.split('$', 1)
    candidate_hash = hashlib.pbkdf2_hmac('sha256', password_candidate.encode('utf-8'), salt.encode('utf-8'), 100000).hex()
    return hmac.compare_digest(candidate_hash, hashed_hex)

def get_avatar_url(name: str, email: str = "") -> str:
    """Generate high-quality avatar URL using UI-Avatars / Gravatar."""
    clean_name = (name or "User").strip().replace(" ", "+")
    color_palette = ["3b82f6", "10b981", "8b5cf6", "ec4899", "f59e0b", "6366f1", "14b8a6"]
    name_hash = sum(ord(c) for c in clean_name) % len(color_palette)
    bg_color = color_palette[name_hash]
    
    ui_avatar = f"https://ui-avatars.com/api/?name={clean_name}&background={bg_color}&color=ffffff&bold=true&size=128&rounded=true"
    
    if email and "@" in email:
        clean_email = email.strip().lower()
        md5_hash = hashlib.md5(clean_email.encode("utf-8")).hexdigest()
        return f"https://www.gravatar.com/avatar/{md5_hash}?s=128&d={ui_avatar}"
    
    return ui_avatar

def get_employees():
    """Retrieve list of all employees."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Danh sách nhân viên"]
    employees = []
    
    for r in range(2, ws.max_row + 1):
        username = ws.cell(row=r, column=1).value
        fullname = ws.cell(row=r, column=2).value
        english_name = ws.cell(row=r, column=3).value
        role = ws.cell(row=r, column=4).value
        email = ws.cell(row=r, column=5).value
        must_change = ws.cell(row=r, column=7).value
        avatar_val = ws.cell(row=r, column=8).value
            
        if username is not None and str(username).strip() != "":
            fn = str(fullname).strip() if fullname else ""
            en = str(english_name).strip() if english_name else ""
            em = str(email).strip() if email else ""
            name_for_avatar = en if en else fn
            avatar_url = str(avatar_val).strip() if avatar_val and str(avatar_val).strip() else get_avatar_url(name_for_avatar, em)
            employees.append({
                "username": str(username).strip(),
                "fullname": fn,
                "english_name": en,
                "role": str(role).strip() if role else "Employee",
                "email": em,
                "must_change_password": bool(must_change),
                "avatar_url": avatar_url
            })
    wb.close()
    return employees

def save_employee(username, fullname, role, english_name="", email="", initial_password=None, avatar_url=None):
    """Add or update an employee."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách nhân viên"]
    
    ws.cell(row=1, column=1).value = 'Username'
    ws.cell(row=1, column=2).value = 'Full Name'
    ws.cell(row=1, column=3).value = 'English Name'
    ws.cell(row=1, column=4).value = 'Role'
    ws.cell(row=1, column=5).value = 'Email'
    ws.cell(row=1, column=6).value = 'PasswordHash'
    ws.cell(row=1, column=7).value = 'MustChangePassword'
    ws.cell(row=1, column=8).value = 'AvatarUrl'
        
    found_row = None
    u_str = str(username).strip()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None and str(val).strip() == u_str:
            found_row = r
            break
            
    pwd_to_use = initial_password if initial_password else DEFAULT_INITIAL_PASSWORD
    pwd_hash = hash_password(pwd_to_use)
    
    if found_row:
        ws.cell(row=found_row, column=2).value = fullname
        ws.cell(row=found_row, column=3).value = english_name
        ws.cell(row=found_row, column=4).value = role
        ws.cell(row=found_row, column=5).value = email
        if avatar_url is not None:
            ws.cell(row=found_row, column=8).value = avatar_url
        if not ws.cell(row=found_row, column=6).value:
            ws.cell(row=found_row, column=6).value = pwd_hash
            ws.cell(row=found_row, column=7).value = True
    else:
        ws.append([u_str, fullname, english_name, role, email, pwd_hash, True, avatar_url or ""])
        
    wb.save(EXCEL_FILE)
    wb.close()

def update_user_avatar(username: str, avatar_url: str):
    """Update user's avatar_url in sheet.xlsx and return updated user object."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách nhân viên"]
    
    u_target = str(username).strip().lower()
    found_row = None
    for r in range(2, ws.max_row + 1):
        u = str(ws.cell(row=r, column=1).value or "").strip()
        if u.lower() == u_target:
            found_row = r
            break
            
    if not found_row:
        wb.close()
        return None
        
    ws.cell(row=found_row, column=8).value = avatar_url.strip()
    wb.save(EXCEL_FILE)
    wb.close()
    
    # Return updated user object
    all_emps = get_employees()
    return next((e for e in all_emps if e["username"].lower() == u_target), None)

def authenticate_user(user_or_email: str, password_candidate: str):
    """Authenticate user by username or email and password."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Danh sách nhân viên"]
    
    query = str(user_or_email).strip().lower()
    candidate_pwd = str(password_candidate or "").strip()
    if not query:
        wb.close()
        return None
        
    found = None
    for r in range(2, ws.max_row + 1):
        u = str(ws.cell(row=r, column=1).value or "").strip()
        fn = str(ws.cell(row=r, column=2).value or "").strip()
        en = str(ws.cell(row=r, column=3).value or "").strip()
        role = str(ws.cell(row=r, column=4).value or "Employee").strip()
        em = str(ws.cell(row=r, column=5).value or "").strip()
        pwd_hash = str(ws.cell(row=r, column=6).value or "").strip()
        must_change = ws.cell(row=r, column=7).value
        avatar_val = ws.cell(row=r, column=8).value
        
        em_prefix = em.split("@")[0].lower() if "@" in em else ""
        if u.lower() == query or (em and em.lower() == query) or (em_prefix and em_prefix == query):
            found = {
                "row": r,
                "username": u,
                "fullname": fn,
                "english_name": en,
                "role": role,
                "email": em,
                "password_hash": pwd_hash,
                "must_change_password": bool(must_change),
                "avatar_url": str(avatar_val).strip() if avatar_val and str(avatar_val).strip() else None
            }
            break
            
    wb.close()
    if not found:
        return None
        
    if verify_password(found["password_hash"], candidate_pwd) or verify_password(found["password_hash"], password_candidate):
        name_for_avatar = found["english_name"] if found["english_name"] else found["fullname"]
        final_avatar = found["avatar_url"] if found["avatar_url"] else get_avatar_url(name_for_avatar, found["email"])
        return {
            "username": found["username"],
            "fullname": found["fullname"],
            "english_name": found["english_name"],
            "role": found["role"],
            "email": found["email"],
            "must_change_password": found["must_change_password"],
            "avatar_url": final_avatar
        }
    return None

def change_user_password(username: str, old_password: str, new_password: str):
    """Change user password after verifying current password."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách nhân viên"]
    
    u_target = str(username).strip().lower()
    found_row = None
    stored_hash = ""
    for r in range(2, ws.max_row + 1):
        u = str(ws.cell(row=r, column=1).value or "").strip()
        if u.lower() == u_target:
            found_row = r
            stored_hash = str(ws.cell(row=r, column=6).value or "").strip()
            break
            
    if not found_row:
        wb.close()
        return False, "Tài khoản không tồn tại."
        
    if not verify_password(stored_hash, old_password):
        wb.close()
        return False, "Mật khẩu hiện tại không đúng."
        
    new_hash = hash_password(new_password)
    ws.cell(row=found_row, column=6).value = new_hash
    ws.cell(row=found_row, column=7).value = False
    
    wb.save(EXCEL_FILE)
    wb.close()
    return True, "Đổi mật khẩu thành công."

def reset_user_password(email: str):
    """Reset user password by registered email and send password reset email via SMTP."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách nhân viên"]
    
    em_target = str(email).strip().lower()
    found_row = None
    found_username = ""
    found_email = ""
    for r in range(2, ws.max_row + 1):
        em = str(ws.cell(row=r, column=5).value or "").strip()
        if em and em.lower() == em_target:
            found_row = r
            found_username = str(ws.cell(row=r, column=1).value or "").strip()
            found_email = em
            break
            
    if not found_row:
        wb.close()
        return False, "Email không tồn tại trong hệ thống. Vui lòng kiểm tra lại.", None
        
    import secrets
    random_pin = secrets.randbelow(899999) + 100000
    temp_password = f"EPM{random_pin}"
    temp_hash = hash_password(temp_password)
    
    ws.cell(row=found_row, column=6).value = temp_hash
    ws.cell(row=found_row, column=7).value = True
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    # Send reset email via SMTP
    try:
        from backend.email_service import send_password_reset_email
        sent_ok, email_msg = send_password_reset_email(found_email, found_username, temp_password)
        if sent_ok:
            return True, f"Mật khẩu mới đã được tự động gửi đến hòm thư {found_email}. Vui lòng kiểm tra hộp thư (hoặc thư rác) để đăng nhập và đổi mật khẩu mới.", temp_password
        else:
            # If SMTP_PASSWORD environment variable is not configured yet, notify user & fallback to temp password info
            return True, f"Đã reset mật khẩu tài khoản thành công!\n({email_msg})\nMật khẩu tạm thời của bạn là: {temp_password}", temp_password
    except Exception as ex:
        return True, f"Mật khẩu của tài khoản '{found_username}' đã được reset về mật khẩu tạm thời: {temp_password}. Vui lòng đăng nhập và đổi mật khẩu mới.", temp_password

def delete_employee(username):
    """Delete employee by username."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Danh sách nhân viên"]
    found_row = None
    u_str = str(username).strip()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None and str(val).strip() == u_str:
            found_row = r
            break
    if found_row:
        ws.delete_rows(found_row)
        # Keep the workbook relationally consistent: an employee removal also
        # removes dependent enrollments and learning progress in the same save.
        for dependent_ws in wb.worksheets:
            is_progress = (
                dependent_ws.cell(1, 1).value == "Username"
                and dependent_ws.cell(1, 5).value == "Status"
            )
            is_enrollment = (
                dependent_ws.cell(1, 1).value == "Username"
                and dependent_ws.cell(1, 2).value == "Target Type"
            )
            if not (is_progress or is_enrollment):
                continue
            for row in range(dependent_ws.max_row, 1, -1):
                value = str(dependent_ws.cell(row=row, column=1).value or "").strip()
                if value == u_str:
                    dependent_ws.delete_rows(row)
        wb.save(EXCEL_FILE)
    wb.close()
    return found_row is not None

def get_tracking_status(status, planned_date_str, completion_date_str, start_date_str=None, duration_minutes=0, planned_start_date_str=None):
    """
    Calculate tracking status based on:
    - Status (Not Started, In Progress, Completed)
    - Planned Completion Date
    - Actual Completion Date / Current Date
    - Start Date & Duration (Short modules 1-3h)
    """
    if not planned_date_str:
        return "On-track"
        
    try:
        if isinstance(planned_date_str, datetime):
            planned_date = planned_date_str.date()
        elif isinstance(planned_date_str, date):
            planned_date = planned_date_str
        else:
            planned_date = datetime.strptime(str(planned_date_str).split()[0], "%Y-%m-%d").date()
    except Exception:
        return "On-track"
        
    today = date.today()
    planned_start_date = None
    if planned_start_date_str:
        try:
            planned_start_date = planned_start_date_str.date() if isinstance(planned_start_date_str, (datetime, date)) else datetime.strptime(str(planned_start_date_str).split()[0], "%Y-%m-%d").date()
        except Exception:
            planned_start_date = None
    
    # Parse start date if available
    start_date = None
    if start_date_str:
        try:
            if isinstance(start_date_str, datetime):
                start_date = start_date_str.date()
            elif isinstance(start_date_str, date):
                start_date = start_date_str
            else:
                start_date = datetime.strptime(str(start_date_str).split()[0], "%Y-%m-%d").date()
        except Exception:
            start_date = None

    is_short_module = (duration_minutes > 0 and duration_minutes <= 180)

    if status == "Completed":
        if completion_date_str:
            try:
                if isinstance(completion_date_str, datetime):
                    comp_date = completion_date_str.date()
                elif isinstance(completion_date_str, date):
                    comp_date = completion_date_str
                else:
                    comp_date = datetime.strptime(str(completion_date_str).split()[0], "%Y-%m-%d").date()
            except Exception:
                comp_date = today
        else:
            comp_date = today
            
        if comp_date <= planned_date:
            return "Fast"
        else:
            return "Slow"
    elif status == "In Progress":
        # Short modules (1-3h) started more than 1 day ago and still incomplete -> Slow
        if is_short_module and start_date and (today - start_date).days > 1:
            return "Slow"

        if today <= planned_date:
            return "On-track"
        else:
            diff_days = (today - planned_date).days
            if diff_days <= 7:
                return "Slow"
            else:
                return "Too slow"
    else:
        # Not Started
        if planned_start_date and today >= planned_start_date:
            return "Too slow" if today > planned_date else "Slow"
        if today < planned_date:
            return "On-track"
        else:
            # If today >= planned_date and user HAS NOT STARTED yet
            diff_days = (today - planned_date).days
            if is_short_module:
                if diff_days <= 3:
                    return "Slow"
                else:
                    return "Too slow"
            else:
                if diff_days <= 7:
                    return "Slow"
                else:
                    return "Too slow"

def _build_enrollment_scopes(courses, enrollments):
    """Return the exact modules each user may see/update from live enrollments."""
    scopes = {}
    for enrollment in enrollments:
        user = str(enrollment.get("username") or "").strip()
        if not user:
            continue
        scope = scopes.setdefault(user, {"module_ids": set(), "module_keys": set(), "enrollment_ids": set()})
        if enrollment.get("enrollment_id"):
            scope["enrollment_ids"].add(enrollment["enrollment_id"])
        target_type = str(enrollment.get("target_type") or "").lower()
        target_id = str(enrollment.get("target_id") or "").strip()
        target_name = str(enrollment.get("target_name") or enrollment.get("course_name") or "").strip()
        for course in courses:
            if course.get("queue") is False:
                continue
            if target_type == "plan":
                matches = (target_id and target_id == course.get("plan_id")) or (not target_id and target_name == course.get("plan"))
            elif target_type == "course":
                matches = (target_id and target_id == course.get("course_id")) or (not target_id and target_name == course.get("course_name"))
            else:
                matches = False
            if matches:
                if course.get("module_id"):
                    scope["module_ids"].add(course["module_id"])
                scope["module_keys"].add((course.get("course_name") or "", course.get("path") or "", course.get("module_name") or ""))
    return scopes


def get_progress(username=None):
    """Retrieve learning progress for one or all users."""
    init_db()
    courses = get_courses()
    module_by_id = {m.get('module_id'): m for m in courses if m.get('module_id')}
    enrollment_scopes = _build_enrollment_scopes(courses, get_enrollments(username))
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Tiến độ học tập"]
    progress_list = []
    
    for r in range(2, ws.max_row + 1):
        u = ws.cell(row=r, column=1).value
        course = ws.cell(row=r, column=2).value
        path = ws.cell(row=r, column=3).value
        module = ws.cell(row=r, column=4).value
        status = ws.cell(row=r, column=5).value
        prog = ws.cell(row=r, column=6).value
        start = ws.cell(row=r, column=7).value
        comp = ws.cell(row=r, column=8).value
        planned = ws.cell(row=r, column=9).value
        track_status = ws.cell(row=r, column=10).value
        planned_start = ws.cell(row=r, column=11).value
        module_id = ws.cell(row=r, column=12).value
        enrollment_id = ws.cell(row=r, column=13).value
        excluded = bool(ws.cell(row=r, column=14).value)
        
        user_key = str(u).strip() if u is not None else ""
        scope = enrollment_scopes.get(user_key)
        module_key = (course or "", path or "", module or "")
        is_enrolled = bool(scope) and (
            (enrollment_id and enrollment_id in scope["enrollment_ids"] and (not module_id or module_id in scope["module_ids"]))
            or (module_id and module_id in scope["module_ids"])
            or module_key in scope["module_keys"]
        )
        if u is not None and not excluded and is_enrolled and (username is None or user_key == str(username).strip()):
            # Format dates to string
            start_str = start.strftime("%Y-%m-%d") if isinstance(start, (datetime, date)) else str(start) if start else ""
            comp_str = comp.strftime("%Y-%m-%d") if isinstance(comp, (datetime, date)) else str(comp) if comp else ""
            planned_str = planned.strftime("%Y-%m-%d") if isinstance(planned, (datetime, date)) else str(planned) if planned else ""
            planned_start_str = planned_start.strftime("%Y-%m-%d") if isinstance(planned_start, (datetime, date)) else str(planned_start) if planned_start else ""
            
            # Recalculate status dynamically in case dates have passed
            course_module = module_by_id.get(module_id)
            track_status = get_tracking_status(status, planned_str, comp_str, start_str, course_module.get('duration_minutes', 0) if course_module else 0, planned_start_str)
            
            progress_list.append({
                "row": r,
                "username": str(u).strip(),
                "course_name": course,
                "path": path if path else "",
                "module_name": module if module else "",
                "status": status if status else "Not Started",
                "progress_percent": float(prog) if prog is not None else 0.0,
                "start_date": start_str,
                "completion_date": comp_str,
                "planned_completion_date": planned_str,
                "tracking_status": track_status,
                "planned_start_date": planned_start_str,
                "module_id": module_id or "",
                "enrollment_id": enrollment_id or ""
            })
    wb.close()
    return progress_list

def save_progress(username, course_name, path, module_name, status, progress_percent, start_date, completion_date, planned_completion_date, module_id=None):
    """Save or update student module progress."""
    init_db()
    courses = get_courses()
    scope = _build_enrollment_scopes(courses, get_enrollments(username)).get(str(username).strip())
    module_key = (course_name or "", path or "", module_name or "")
    if not scope or not ((module_id and module_id in scope["module_ids"]) or module_key in scope["module_keys"]):
        raise ValueError("No Course assigned. Register the user before updating progress.")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Tiến độ học tập"]
    
    # Normalize status and percent alignment
    try:
        progress_percent = float(progress_percent)
    except Exception:
        progress_percent = 0.0

    if progress_percent >= 100.0:
        status = "Completed"
        progress_percent = 100.0
    elif progress_percent > 0.0 and status == "Not Started":
        status = "In Progress"

    today_val = date.today()
    if status in ("In Progress", "Completed") and not start_date:
        start_val = today_val
    elif start_date:
        try:
            start_val = datetime.strptime(str(start_date).split()[0], "%Y-%m-%d").date()
        except Exception:
            start_val = today_val
    else:
        start_val = None

    if status == "Completed":
        if not completion_date:
            comp_val = today_val
        else:
            try:
                comp_val = datetime.strptime(str(completion_date).split()[0], "%Y-%m-%d").date()
            except Exception:
                comp_val = today_val
    else:
        comp_val = None

    if planned_completion_date:
        try:
            planned_val = datetime.strptime(str(planned_completion_date).split()[0], "%Y-%m-%d").date()
        except Exception:
            planned_val = None
    else:
        planned_val = None

    # Calculate status
    tracking = get_tracking_status(status, planned_val, comp_val)

    found_row = None
    for r in range(2, ws.max_row + 1):
        u = ws.cell(row=r, column=1).value
        c = ws.cell(row=r, column=2).value
        p = ws.cell(row=r, column=3).value
        m = ws.cell(row=r, column=4).value
        stored_module_id = ws.cell(row=r, column=12).value
        
        # Check matching record
        identity_matches = (stored_module_id == module_id) if module_id else (c == course_name and (p or "") == (path or "") and m == module_name)
        if u is not None and str(u).strip() == str(username).strip() and identity_matches:
            found_row = r
            break

    u_str = str(username).strip()
    if found_row:
        ws.cell(row=found_row, column=5).value = status
        ws.cell(row=found_row, column=6).value = float(progress_percent)
        ws.cell(row=found_row, column=7).value = start_val
        ws.cell(row=found_row, column=8).value = comp_val
        ws.cell(row=found_row, column=9).value = planned_val
        ws.cell(row=found_row, column=10).value = tracking
        if module_id:
            ws.cell(row=found_row, column=12).value = module_id
    else:
        ws.append([
            u_str, course_name, path or "", module_name, status, 
            float(progress_percent), start_val, comp_val, 
            planned_val, tracking, None, module_id or "", "", False
        ])
        
    wb.save(EXCEL_FILE)
    wb.close()


def get_working_days(start_date_str, end_date_str, workweek_type):
    """
    Generate list of YYYY-MM-DD date strings between start and end (inclusive)
    excluding weekends depending on workweek_type:
    - 5: exclude Saturday (5) and Sunday (6)
    - 6: exclude Sunday (6)
    - 7: include all days
    """
    from datetime import timedelta
    try:
        start_date = datetime.strptime(str(start_date_str).split()[0], "%Y-%m-%d").date()
        end_date = datetime.strptime(str(end_date_str).split()[0], "%Y-%m-%d").date()
    except Exception:
        return []
        
    working_days = []
    curr = start_date
    while curr <= end_date:
        weekday = curr.weekday() # 0 = Monday, ..., 5 = Saturday, 6 = Sunday
        if workweek_type == 5:
            if weekday < 5: # Monday to Friday
                working_days.append(curr.strftime("%Y-%m-%d"))
        elif workweek_type == 6:
            if weekday < 6: # Monday to Saturday
                working_days.append(curr.strftime("%Y-%m-%d"))
        elif workweek_type == 7:
            working_days.append(curr.strftime("%Y-%m-%d"))
        curr += timedelta(days=1)
    return working_days


def get_enrollments(username=None):
    """Get list of course/plan enrollments."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Đăng ký khóa học"]
    enrollments = []
    
    has_target_type = (ws.cell(row=1, column=2).value == 'Target Type')
    
    for r in range(2, ws.max_row + 1):
        u = ws.cell(row=r, column=1).value
        if has_target_type:
            target_type = ws.cell(row=r, column=2).value or "plan"
            target_name = ws.cell(row=r, column=3).value or ""
            start = ws.cell(row=r, column=4).value
            end = ws.cell(row=r, column=5).value
            ww = ws.cell(row=r, column=6).value
            ratio_val = ws.cell(row=r, column=7).value
            daily_h = ws.cell(row=r, column=8).value
            enrollment_id = ws.cell(row=r, column=9).value
            target_id = ws.cell(row=r, column=10).value
            target_version = ws.cell(row=r, column=11).value
        else:
            target_type = "course"
            target_name = ws.cell(row=r, column=2).value or ""
            start = ws.cell(row=r, column=3).value
            end = ws.cell(row=r, column=4).value
            ww = ws.cell(row=r, column=5).value
            ratio_val = 3.0
            daily_h = 2.0
            enrollment_id = ""
            target_id = ""
            target_version = 1
            
        if u is not None and str(u).strip() != "":
            start_str = start.strftime("%Y-%m-%d") if isinstance(start, (datetime, date)) else str(start).split()[0] if start else ""
            end_str = end.strftime("%Y-%m-%d") if isinstance(end, (datetime, date)) else str(end).split()[0] if end else ""
            
            if username is None or str(u).strip() == str(username).strip():
                enrollments.append({
                    "username": str(u).strip(),
                    "target_type": str(target_type).lower(),
                    "target_name": target_name,
                    "course_name": target_name,
                    "start_date": start_str,
                    "planned_end_date": end_str,
                    "workweek_type": int(ww) if ww is not None else 5,
                    "ratio": float(ratio_val) if ratio_val is not None else 3.0,
                    "daily_hours": float(daily_h) if daily_h is not None else 2.0,
                    "enrollment_id": enrollment_id or _stable_id('enr', u, target_type, target_name),
                    "target_id": target_id or "",
                    "target_version": int(target_version or 1)
                })
    wb.close()
    return enrollments

def _is_working_day(day_value, workweek_type):
    weekday = day_value.weekday()
    return workweek_type == 7 or (workweek_type == 6 and weekday < 6) or (workweek_type == 5 and weekday < 5)

def _working_day_at(start_value, index, workweek_type):
    current = start_value
    while not _is_working_day(current, workweek_type):
        current += timedelta(days=1)
    seen = 0
    while seen < index:
        current += timedelta(days=1)
        if _is_working_day(current, workweek_type):
            seen += 1
    return current

def _effective_minutes(module, ratio):
    is_exam = bool(re.search(r'1z0-|exam|certification|professional', module.get('module_name', ''), re.I))
    base = max(0, int(module.get('duration_minutes') or 0))
    return max(1, int(round(base if is_exam else base * float(ratio))))

def _get_plan_meta_by_name(plan_name):
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[PLAN_META_SHEET]
    result = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 2).value or '').strip() == str(plan_name or '').strip():
            result = {'plan_id': ws.cell(r, 1).value, 'plan_type': ws.cell(r, 3).value or 'main', 'source_plan_id': ws.cell(r, 4).value or '', 'version': int(ws.cell(r, 6).value or 1)}
            break
    wb.close()
    return result or {'plan_id': _stable_id('plan', plan_name), 'plan_type': 'main', 'source_plan_id': '', 'version': 1}

def _schedule_enrollment(username, target_type, target_name, start_date, workweek_type, ratio, daily_hours, enrollment_id=None, target_id=None):
    """Schedule modules sequentially using the enrollment's configured study-day capacity."""
    all_courses = get_courses()
    if str(target_type).lower() == 'plan':
        target_catalog = [m for m in all_courses if (target_id and m.get('plan_id') == target_id) or (not target_id and m['plan'] == target_name)]
        modules = [m for m in target_catalog if m.get('queue', True)]
    else:
        target_catalog = [m for m in all_courses if (target_id and m.get('course_id') == target_id) or (not target_id and m['course_name'] == target_name)]
        modules = [m for m in target_catalog if m.get('queue', True)]
    modules.sort(key=lambda m: (m.get('course_order', 0), m.get('module_order', 0), m.get('row', 0)))
    if not modules:
        return start_date

    u_str = str(username).strip()
    enrollment_id = enrollment_id or _stable_id('enr', u_str, target_type, target_name)
    start_val = datetime.strptime(str(start_date).split()[0], '%Y-%m-%d').date()
    ww = int(workweek_type)
    day_capacity_minutes = max(30, int(round(float(daily_hours or 2.0) * 60)))
    schedule = []
    cursor = 0
    for module in modules:
        minutes = _effective_minutes(module, ratio)
        start_idx = cursor // day_capacity_minutes
        end_idx = max(start_idx, (cursor + minutes - 1) // day_capacity_minutes)
        start_offset = cursor % day_capacity_minutes
        end_cursor = cursor + minutes
        end_offset = end_cursor % day_capacity_minutes or day_capacity_minutes
        schedule.append((module, _working_day_at(start_val, start_idx, ww), _working_day_at(start_val, end_idx, ww), start_offset, end_offset))
        cursor += minutes

    wb = openpyxl.load_workbook(EXCEL_FILE)
    p_ws = next(s for s in wb.worksheets if s.cell(1, 1).value == 'Username' and s.cell(1, 5).value == 'Status')
    active_ids = {m.get('module_id') for m, _, _, _, _ in schedule}
    catalog_by_key = {(m['course_name'], m['path'] or '', m['module_name']): m for m in target_catalog}
    existing = {}
    for r in range(2, p_ws.max_row + 1):
        if str(p_ws.cell(r, 1).value or '').strip() != u_str:
            continue
        row_enrollment = p_ws.cell(r, 13).value
        module_id = p_ws.cell(r, 12).value
        legacy_key = (p_ws.cell(r, 2).value, p_ws.cell(r, 3).value or '', p_ws.cell(r, 4).value)
        catalog_module = catalog_by_key.get(legacy_key)
        if not row_enrollment and catalog_module and not catalog_module.get('queue') and (p_ws.cell(r, 5).value or 'Not Started') == 'Not Started':
            p_ws.cell(r, 12).value = catalog_module.get('module_id')
            p_ws.cell(r, 13).value = enrollment_id
            p_ws.cell(r, 14).value = True
        if row_enrollment == enrollment_id and module_id:
            existing[module_id] = r
            if module_id not in active_ids and (p_ws.cell(r, 5).value or 'Not Started') == 'Not Started':
                p_ws.cell(r, 14).value = True

    for module, planned_start, planned_end, start_offset, end_offset in schedule:
        module_id = module.get('module_id')
        row = existing.get(module_id)
        if row is None:
            # Adopt a legacy row once, preserving its history.
            for r in range(2, p_ws.max_row + 1):
                if (str(p_ws.cell(r, 1).value or '').strip() == u_str and p_ws.cell(r, 2).value == module['course_name'] and (p_ws.cell(r, 3).value or '') == (module['path'] or '') and p_ws.cell(r, 4).value == module['module_name'] and not p_ws.cell(r, 13).value):
                    row = r
                    break
        if row is None:
            p_ws.append([u_str, module['course_name'], module['path'] or '', module['module_name'], 'Not Started', 0.0, None, None, planned_end, '', planned_start, module_id, enrollment_id, False, start_offset, end_offset])
            row = p_ws.max_row
        else:
            p_ws.cell(row, 12).value = module_id
            p_ws.cell(row, 13).value = enrollment_id
            p_ws.cell(row, 14).value = False
            # Planned dates describe the current plan version. Actual dates are
            # kept separately and remain untouched as historical evidence.
            p_ws.cell(row, 9).value = planned_end
            p_ws.cell(row, 11).value = planned_start
        p_ws.cell(row, 15).value = start_offset
        p_ws.cell(row, 16).value = end_offset
        p_ws.cell(row, 10).value = get_tracking_status(p_ws.cell(row, 5).value or 'Not Started', p_ws.cell(row, 9).value, p_ws.cell(row, 8).value, p_ws.cell(row, 7).value, module.get('duration_minutes', 0), p_ws.cell(row, 11).value)
    wb.save(EXCEL_FILE)
    wb.close()
    return schedule[-1][2].strftime('%Y-%m-%d')

def _touch_plan_and_reschedule(plan_name):
    """Increment a Plan version and reschedule every live enrollment mapped to it."""
    if not plan_name or not os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if PLAN_META_SHEET not in wb.sheetnames:
        wb.close()
        return
    meta_ws = wb[PLAN_META_SHEET]
    version = 1
    plan_id = ''
    for r in range(2, meta_ws.max_row + 1):
        if str(meta_ws.cell(r, 2).value or '').strip() == str(plan_name).strip():
            plan_id = meta_ws.cell(r, 1).value
            version = int(meta_ws.cell(r, 6).value or 1) + 1
            meta_ws.cell(r, 6).value = version
            meta_ws.cell(r, 7).value = datetime.now()
            break
    wb.save(EXCEL_FILE)
    wb.close()
    for enrollment in get_enrollments():
        if enrollment['target_type'] == 'plan' and (enrollment['target_name'] == plan_name or (plan_id and enrollment.get('target_id') == plan_id)):
            end_date = _schedule_enrollment(enrollment['username'], enrollment['target_type'], enrollment['target_name'], enrollment['start_date'], enrollment['workweek_type'], enrollment['ratio'], enrollment['daily_hours'], enrollment.get('enrollment_id'), enrollment.get('target_id'))
            _update_enrollment_schedule_metadata(enrollment.get('enrollment_id'), end_date, plan_id, version)

def _update_enrollment_schedule_metadata(enrollment_id, end_date, target_id, target_version):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = next((s for s in wb.worksheets if s.cell(1, 2).value == 'Target Type'), None)
    if ws:
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 9).value == enrollment_id:
                ws.cell(r, 5).value = datetime.strptime(end_date, '%Y-%m-%d').date()
                ws.cell(r, 10).value = target_id
                ws.cell(r, 11).value = target_version
                break
        wb.save(EXCEL_FILE)
    wb.close()


def save_enrollment(username, target_type, target_name, start_date, end_date, workweek_type, ratio=3.0, daily_hours=2.0, target_id=None):
    """Save enrollment (Plan or Course) and auto-schedule module deadlines."""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Đăng ký khóa học"]
    
    if ws.cell(row=1, column=2).value != 'Target Type':
        ws.cell(row=1, column=2).value = 'Target Type'
        ws.cell(row=1, column=3).value = 'Target Name'
        ws.cell(row=1, column=4).value = 'Start Date'
        ws.cell(row=1, column=5).value = 'Planned End Date'
        ws.cell(row=1, column=6).value = 'Workweek Type'
        ws.cell(row=1, column=7).value = 'Ratio'
        ws.cell(row=1, column=8).value = 'Daily Hours'
        ws.cell(row=1, column=9).value = 'Enrollment ID'
        ws.cell(row=1, column=10).value = 'Target ID'
        ws.cell(row=1, column=11).value = 'Target Version'
        
    found_row = None
    u_str = str(username).strip()
    for r in range(2, ws.max_row + 1):
        u = ws.cell(row=r, column=1).value
        t_name = ws.cell(row=r, column=3).value if ws.cell(row=1, column=2).value == 'Target Type' else ws.cell(row=r, column=2).value
        row_target_type = str(ws.cell(r, 2).value or 'plan').lower()
        row_target_id = str(ws.cell(r, 10).value or '')
        target_matches = (row_target_id == str(target_id)) if target_id else (t_name == target_name)
        if u is not None and str(u).strip() == u_str and target_matches and row_target_type == str(target_type).lower():
            found_row = r
            break
            
    start_val = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_val = datetime.strptime(end_date, "%Y-%m-%d").date()
    enrollment_id = _stable_id('enr', u_str, target_type, target_name)
    target_meta = _get_plan_meta_by_name(target_name) if str(target_type).lower() == 'plan' else {'plan_id': target_id or '', 'version': 1}
    resolved_target_id = target_id or target_meta.get('plan_id', '')
    
    if found_row:
        ws.cell(row=found_row, column=2).value = target_type
        ws.cell(row=found_row, column=3).value = target_name
        ws.cell(row=found_row, column=4).value = start_val
        ws.cell(row=found_row, column=5).value = end_val
        ws.cell(row=found_row, column=6).value = int(workweek_type)
        ws.cell(row=found_row, column=7).value = float(ratio)
        ws.cell(row=found_row, column=8).value = float(daily_hours)
        ws.cell(row=found_row, column=9).value = enrollment_id
        ws.cell(row=found_row, column=10).value = resolved_target_id
        ws.cell(row=found_row, column=11).value = target_meta.get('version', 1)
    else:
        ws.append([u_str, target_type, target_name, start_val, end_val, int(workweek_type), float(ratio), float(daily_hours), enrollment_id, resolved_target_id, target_meta.get('version', 1)])
        
    wb.save(EXCEL_FILE)
    wb.close()

    computed_end = _schedule_enrollment(u_str, target_type, target_name, start_date, workweek_type, ratio, daily_hours, enrollment_id, resolved_target_id)
    _update_enrollment_schedule_metadata(enrollment_id, computed_end, resolved_target_id, target_meta.get('version', 1))
    return
    
    # 2. Schedule module deadlines
    all_courses = get_courses()
    if target_type.lower() == "plan":
        target_modules = [m for m in all_courses if m["plan"] == target_name and m.get("queue", True)]
    else:
        target_modules = [m for m in all_courses if m["course_name"] == target_name and m.get("queue", True)]
        
    if not target_modules:
        return
        
    working_days = get_working_days(start_date, end_date, int(workweek_type))
    N = len(working_days)
    if N == 0:
        working_days = [end_date]
        N = 1
        
    import re
    def get_effective_mins(m, r_val):
        m_name = m.get("module_name", "")
        is_exam = bool(re.search(r'1z0-|exam|certification|professional', m_name, re.I))
        dm = m.get("duration_minutes", 0)
        return dm if is_exam else dm * r_val

    r_factor = float(ratio)
    total_effective_mins = sum(get_effective_mins(m, r_factor) for m in target_modules)
    if total_effective_mins == 0:
        total_effective_mins = len(target_modules)
        
    cum_effective_mins = 0
    progress_sheet = openpyxl.load_workbook(EXCEL_FILE)
    p_ws = progress_sheet["Tiến độ học tập"]
    
    for m in target_modules:
        cum_effective_mins += get_effective_mins(m, r_factor)
        weight = cum_effective_mins / total_effective_mins
        idx = int(round(weight * (N - 1)))
        assigned_date = working_days[idx]
        
        existing_row = None
        for r in range(2, p_ws.max_row + 1):
            pu = p_ws.cell(row=r, column=1).value
            pc = p_ws.cell(row=r, column=2).value
            pp = p_ws.cell(row=r, column=3).value
            pm = p_ws.cell(row=r, column=4).value
            if pu is not None and str(pu).strip() == u_str and pc == m["course_name"] and (pp or "") == (m["path"] or "") and pm == m["module_name"]:
                existing_row = r
                break
                
        tracking = get_tracking_status(
            "Not Started" if not existing_row else p_ws.cell(row=existing_row, column=5).value,
            assigned_date,
            None if not existing_row else p_ws.cell(row=existing_row, column=8).value
        )
        
        if existing_row:
            p_ws.cell(row=existing_row, column=9).value = datetime.strptime(assigned_date, "%Y-%m-%d").date()
            p_ws.cell(row=existing_row, column=10).value = tracking
        else:
            p_ws.append([
                u_str, m["course_name"], m["path"] or "", m["module_name"],
                "Not Started", 0.0, None, None, 
                datetime.strptime(assigned_date, "%Y-%m-%d").date(), tracking
            ])
            
    progress_sheet.save(EXCEL_FILE)
    progress_sheet.close()


def delete_enrollment(username, target_name):
    """Delete enrollment and related learning progress modules."""
    init_db()
    u_str = str(username).strip()
    lookup_wb = openpyxl.load_workbook(EXCEL_FILE)
    enrollment_ws = next((s for s in lookup_wb.worksheets if s.cell(1, 2).value == 'Target Type'), None)
    enrollment_row = None
    enrollment_id = None
    if enrollment_ws:
        for r in range(2, enrollment_ws.max_row + 1):
            if str(enrollment_ws.cell(r, 1).value or '').strip() == u_str and enrollment_ws.cell(r, 3).value == target_name:
                enrollment_row = r
                enrollment_id = enrollment_ws.cell(r, 9).value
                break
    if enrollment_row and enrollment_id:
        progress_ws = next(s for s in lookup_wb.worksheets if s.cell(1, 5).value == 'Status')
        for r in range(progress_ws.max_row, 1, -1):
            if progress_ws.cell(r, 13).value == enrollment_id:
                progress_ws.delete_rows(r)
        enrollment_ws.delete_rows(enrollment_row)
        lookup_wb.save(EXCEL_FILE)
        lookup_wb.close()
        return
    lookup_wb.close()
    
    all_courses = get_courses()
    target_modules = [m for m in all_courses if m["plan"] == target_name or m["course_name"] == target_name]
    
    wb_prog = openpyxl.load_workbook(EXCEL_FILE)
    ws_p = wb_prog["Tiến độ học tập"]
    rows_to_delete_p = []
    
    for r in range(2, ws_p.max_row + 1):
        u = ws_p.cell(row=r, column=1).value
        c = ws_p.cell(row=r, column=2).value
        if u is not None and str(u).strip() == u_str:
            if c == target_name or any(tm["course_name"] == c for tm in target_modules):
                rows_to_delete_p.append(r)
                
    for r in sorted(rows_to_delete_p, reverse=True):
        ws_p.delete_rows(r)
    wb_prog.save(EXCEL_FILE)
    wb_prog.close()
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Đăng ký khóa học"]
    found_row = None
    has_target_type = (ws.cell(row=1, column=2).value == 'Target Type')
    name_col = 3 if has_target_type else 2
    
    for r in range(2, ws.max_row + 1):
        u = ws.cell(row=r, column=1).value
        name_val = ws.cell(row=r, column=name_col).value
        if u is not None and str(u).strip() == u_str and name_val == target_name:
            found_row = r
            break
            
    if found_row:
        ws.delete_rows(found_row)
        wb.save(EXCEL_FILE)
    wb.close()

